from django.shortcuts import render,redirect,get_object_or_404
from django.contrib import messages
from .models import *
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from cart_management.models import OrderDetails
from django.contrib.auth.decorators import user_passes_test
from coupon.models import Coupon
from banner.models import Banner
from reportlab.pdfgen import canvas
from django.db.models import Sum
from datetime import datetime,timedelta 
from django.db.models import Count
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils import timezone
from django.db.models.functions import TruncWeek, TruncMonth, TruncYear
from django.db.models.functions import TruncDate
from offer.models import ProductOffer
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

# Create your views here.
# By using this only admin can access these pages
def is_superuser(user):
    return user.is_authenticated and user.is_superuser

@user_passes_test(is_superuser, login_url='home:signin')
def admin_dashboard(request):

    #To find the revenue 
    total_revenue = OrderDetails.objects.aggregate(total_amount=Sum('payment__amount'))
    total_revenue_amount = total_revenue['total_amount']
    if total_revenue_amount is None:
        total_revenue_amount = 0

    #To find total number of orders
    total_orders = OrderDetails.objects.count()

    # Total number of products
    total_products = Product.objects.count()
    # Total number of categories
    total_categories = Category.objects.count()

    #Monthly revernue
    first_day = datetime(2024, 1, 1)
    last_day = datetime(2024, 2 + 1, 1) - timedelta(days=1)
    # Use aggregate to sum up the total amount from payments within the given month
    monthly_earning = Payment.objects.filter(payment_date__range=(first_day, last_day)).aggregate(total_amount=Sum('amount'))
    # Access the total earning value
    total_earning_amount = monthly_earning['total_amount']
    # If there are no earnings for the month, handle the case where total_earning_amount is None
    if total_earning_amount is None:
        total_earning_amount = 0

    #Total number of users
    total_user = User.objects.count()

    #To get all top selling product varient
    # Perform aggregation to get the count of orders for each product variant
    top_selling_variants = OrderDetails.objects.values('product_variant').annotate(order_count=Count('id')).order_by('-order_count')[:10]
    # Now you have the top 10 selling product variants, and you can retrieve the corresponding ProductVariant instances
    top_selling_product_variants = ProductVariant.objects.filter(id__in=[item['product_variant'] for item in top_selling_variants])
    # If you need more information about each variant, you can prefetch related data
    top_selling_product_variants = top_selling_product_variants.prefetch_related('product', 'product__category')

    #To get top selling categories
    top_categories = Category.objects.annotate(order_count=Count('products__orderdetails')).order_by('-order_count')[:10]

    #To get latest created user
    latest_users = User.objects.order_by('-id')[:10]


    #report chart

    # Assuming you have a model named OrderDetails
    weekly_data = OrderDetails.objects.annotate(week=TruncWeek('order_time')).values('week').annotate(order_count=Count('id'))
    monthly_data = OrderDetails.objects.annotate(month=TruncMonth('order_time')).values('month').annotate(order_count=Count('id'))
    yearly_data = OrderDetails.objects.annotate(year=TruncYear('order_time')).values('year').annotate(order_count=Count('id'))

    # Convert QuerySet data to lists for chart
    weekly_labels = [entry['week'].strftime('%Y-%m-%d') for entry in weekly_data]
    weekly_counts = [entry['order_count'] for entry in weekly_data]

    monthly_labels = [entry['month'].strftime('%Y-%m') for entry in monthly_data]
    monthly_counts = [entry['order_count'] for entry in monthly_data]

    yearly_labels = [entry['year'].strftime('%Y') for entry in yearly_data]
    yearly_counts = [entry['order_count'] for entry in yearly_data]

    print(weekly_labels)
    print(weekly_counts)

    print(monthly_labels)
    print(monthly_counts)

    print(yearly_labels)
    print(yearly_counts)

    today = timezone.now().date()
    start_of_week = today - timezone.timedelta(days=today.weekday())
    end_of_week = start_of_week + timezone.timedelta(days=6)

    # Query to get the number of sales with status 'Delivered' for each day of the week
    weekly_sales = OrderDetails.objects.filter(
        order_status='Delivered',
        order_time__date__range=[start_of_week, end_of_week]
    ).values('order_time__date').annotate(total_sales=Count('id')).order_by('order_time__date')

    # Create a list of 7 integers representing the number of sales for each day of the week
    weekly_sales_list = [0] * 7
    for sale in weekly_sales:
        day_of_week = (sale['order_time__date'] - start_of_week).days
        weekly_sales_list[day_of_week] = sale['total_sales']

    print(weekly_sales_list)


    # Get the current month's start and end dates
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    end_of_month = (start_of_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)

    # Query to get the number of sales with status 'Delivered' for each day of the month
    monthly_sales = OrderDetails.objects.filter(
        order_status='Delivered',
        order_time__date__range=[start_of_month, end_of_month]
    ).values('order_time__date').annotate(total_sales=Count('id')).order_by('order_time__date')

    # Create a list of integers representing the number of sales for each day of the month
    num_days_in_month = (end_of_month - start_of_month).days + 1
    monthly_sales_list = [0] * num_days_in_month
    for sale in monthly_sales:
        day_of_month = (sale['order_time__date'] - start_of_month).days
        monthly_sales_list[day_of_month] = sale['total_sales']

    print(monthly_sales_list)


    start_of_year = timezone.now().replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_year = start_of_year.replace(year=start_of_year.year + 1) - timezone.timedelta(days=1)

    yearly_sales = OrderDetails.objects.filter(
        order_status='Delivered',
        date__range=[start_of_year, end_of_year]
    ).values('date__month').annotate(total_sales=Count('id')).order_by('date__month')

    yearly_sales_list = [0] * 12
    for sale in yearly_sales:
        month_index = sale['date__month'] - 1
        yearly_sales_list[month_index] = sale['total_sales']


    print(yearly_sales_list)
    return render(request,'adminpage/admin_main.html',{'revenue':total_revenue_amount,
                                                       'total_orders':total_orders,
                                                       'tp':total_products,
                                                       'tc':total_categories,
                                                       'tu':total_user,
                                                       'topp':top_selling_product_variants  ,
                                                       'topc':top_categories ,
                                                       'ls':latest_users,
                                                       'a':weekly_sales_list,  
                                                       'b':monthly_sales_list, 
                                                       'c':yearly_sales_list,                                       
                                                       })

@user_passes_test(is_superuser, login_url='home:signin')
def user_management(request):
    users = User.objects.exclude(is_superuser=True)
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number)
    return render(request, 'adminpage/user_management.html', {'users': users_page})
    
@user_passes_test(is_superuser, login_url='home:signin')
def category_management(request):
    categories = Category.objects.all()
    paginator = Paginator(categories, 5)
    page_number = request.GET.get('page')
    categories_page = paginator.get_page(page_number)
    return render(request, 'adminpage/category_list.html', {'categories': categories_page})

@user_passes_test(is_superuser, login_url='home:signin')
def product_management(request):
    products = Product.objects.all()
    paginator = Paginator(products, 5)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    return render(request, 'adminpage/product_list.html', {'products': products_page})

@user_passes_test(is_superuser, login_url='home:signin')
def edit_product(request):
    return render(request,'adminpage/edit_product.html')

@user_passes_test(is_superuser, login_url='home:signin')
def update_category(request):
    if request.method == "POST": 
        category_name = request.POST.get("name")
        category_description = request.POST.get("description")
        if Category.objects.filter(name=category_name).exists():
            return redirect('adminpage:category_management')
        category=Category.objects.create(name=category_name,description=category_description)
        print(category)
        category.save()
        return redirect('adminpage:category_management')
    else:
        return render(request,'adminpage/update_category.html')
    
@user_passes_test(is_superuser, login_url='home:signin') 
def delete_category(request,id):
    category=Category.objects.get(id=id)
    if category.is_active==True:
        category.is_active=False
    else:
        category.is_active=True
    category.save()
    return redirect('adminpage:category_management')

@user_passes_test(is_superuser, login_url='home:signin')
def category_update(request,id):
    category=Category.objects.get(id=id)
    if request.method == 'POST':
        name=request.POST.get('name')
        description=request.POST.get('description')
        if name:
            category.name=name
        if description:
            category.description=description
        category.save()
        return redirect('adminpage:category_management')
    else:
        return render(request, 'adminpage/category_update.html',{'cu':category})

@user_passes_test(is_superuser, login_url='home:signin')
def add_products(request):
    if request.method == 'POST':
        # Extract product details
        title = request.POST.get('title').strip()
        description = request.POST.get('description').strip()
        category_name = request.POST.get('category')

        # Validations for title and description
        if not title or not description:
            messages.error(request, 'Title and description cannot be empty or spaces.')
            return redirect('adminpage:add_products')  # Updated redirect URL

        # Handle main image upload
        main_image = request.FILES.get('image') if 'image' in request.FILES else None

        # Handle multiple images upload
        additional_images = request.FILES.getlist('images')

        # Assuming you have a Categories model with a name field
        category, created = Category.objects.get_or_create(name=category_name)

        # Get an existing product or create a new one
        product, created = Product.objects.get_or_create(
            category=category,
            title=title,
            defaults={
                'description': description,
                'image': main_image,
            }
        )

        # Create MultipleImage instances for additional images
        for img in additional_images:
            multipleImage.objects.create(product=product, image=img)

        # Extract variant details
        size = request.POST.get('size')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')

        # Validations for size, quantity, and price
        if not size or not price or not quantity:
            messages.error(request, 'Size, price, and quantity cannot be empty.')
            return redirect('adminpage:add_products')  # Updated redirect URL

        try:
            size = int(size)
            price = float(price)
            quantity = int(quantity)
        except ValueError:
            messages.error(request, 'Size and quantity should be integers. Price should be a valid number.')
            return redirect('adminpage:add_products')  # Updated redirect URL

        if size < 1 or price < 1 or quantity < 1:
            messages.error(request, 'Size, price, and quantity should be greater than or equal to 1.')
            return redirect('adminpage:add_products')  # Updated redirect URL

        # Create the product variant
        ProductVariant.objects.create(
            product=product,
            size=size,
            price=price,
            quantity=quantity
        )

        return redirect('adminpage:product_management')


    # Fetch categories from the database
    categories = Category.objects.all()
    return render(request, 'adminpage/add_products.html', {'categories': categories})

@user_passes_test(is_superuser, login_url='home:signin')
def update_product(request,id):
    # Get the product instance using the product_id
    product = get_object_or_404(Product, id=id)
    if request.method == 'POST':
        # Process the form submission
        title = request.POST.get('title')
        description = request.POST.get('description')
        price = request.POST.get('price')
        image = request.FILES.get('image') if 'image' in request.FILES else None
        # Update the product fields
        product.title = title
        product.description = description
        product.price = price

        if image:
            product.image = image

        product.save()    

        return redirect('adminpage:product_management')  # Redirect to the provider panel or any other desired page
    return render(request, 'adminpage/edit_product.html', {'product': product})

@user_passes_test(is_superuser, login_url='home:signin')
def activate_deactivate(request,id=id):
    if request.method=='POST':
        ban=Product.objects.get(id=id)
        if ban.is_active==True:
            ban.is_active=False
        else:
            ban.is_active=True
        ban.save()
    return redirect('adminpage:product_management')

@user_passes_test(is_superuser, login_url='home:signin')
def activate_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    return redirect('adminpage:user_management')  # Redirect to the user list or any other page

@user_passes_test(is_superuser, login_url='home:signin')
def deactivate_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = False
    user.save()
    return redirect('adminpage:user_management')  # Redirect to the user list or any other page

@user_passes_test(is_superuser, login_url='home:signin')
def order_list(request):
    orders = OrderDetails.objects.all().order_by('-id')
    paginator = Paginator(orders, 5)
    page_number = request.GET.get('page')
    try:
        orders_page = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        orders_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        orders_page = paginator.page(paginator.num_pages)
    return render(request, 'adminpage/order_list.html', {'orders': orders_page})

@user_passes_test(is_superuser, login_url='home:signin')
def varients(request,id):
    product = get_object_or_404(Product, id=id)
    return render(request,'adminpage/varients.html',{'product':product})

@user_passes_test(is_superuser, login_url='home:signin')
def add_varients(request, id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=id)
        size = request.POST.get('size')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')

        # Validate that size is not empty and not negative
        if not size or float(size) < 0:
            error_message = 'Please enter a valid size.'
            messages.error(request, error_message)
            return redirect('adminpage:add_varients', id=id)
        
        # Validate that price is a valid number
        try:
            price = float(price)
            if price <= 0:
                raise ValueError('Price must be greater than 0.')
        except ValueError:
            error_message = 'Please enter a valid price.'
            messages.error(request, error_message)
            return redirect('adminpage:add_varients', id=id)

        # Validate that quantity is an integer and not less than 1
        try:
            quantity = int(quantity)
            if quantity < 1:
                raise ValueError('Quantity must be 1 or greater.')
        except ValueError:
            error_message = 'Please enter a valid quantity.'
            messages.error(request, error_message)
            return redirect('adminpage:add_varients', id=id)

        variant = ProductVariant(product=product, size=size, price=price, quantity=quantity)

        # Save the variant
        variant.save()

        # Redirect to some success page or back to the product details page
        return redirect('adminpage:varients', id=id)

    return render(request, 'adminpage/add_varients.html')

@user_passes_test(is_superuser, login_url='home:signin')
def cancel_order(request,id):
    print(id)
    product = OrderDetails.objects.get(id = id)
    product.order_status = 'Cancelled'
    product.save()
    return redirect('adminpage:order_list')

@user_passes_test(is_superuser, login_url='home:signin')
def order_delivered(request,id):
    print(id)
    product = OrderDetails.objects.get(id = id)
    product.order_status = 'Delivered'
    product.save()
    return redirect('adminpage:order_list')

@user_passes_test(is_superuser, login_url='home:signin')
def refund_order(request,id):
    print(id)
    product = OrderDetails.objects.get(id = id)
    product.order_status = 'Refund'
    product.save()
    return redirect('adminpage:order_list')

@user_passes_test(is_superuser, login_url='home:signin')
def replacement_order(request,id):
    print(id)
    product = OrderDetails.objects.get(id = id)
    product.order_status = 'Replacement'
    product.save()
    return redirect('adminpage:order_list')

@user_passes_test(is_superuser, login_url='home:signin')
def return_order(request,id):
    print(id)
    product = OrderDetails.objects.get(id = id)
    product.order_status = 'Return'
    product.save()
    return redirect('adminpage:order_list')

@user_passes_test(is_superuser, login_url='home:signin')
def coupon_management(request):
    coupons = Coupon.objects.all()
    paginator = Paginator(coupons, 5)
    page_number = request.GET.get('page')
    coupons_page = paginator.get_page(page_number)
    return render(request, 'adminpage/coupon_management.html', {'coupons': coupons_page})
    
@user_passes_test(is_superuser, login_url='home:signin')
def add_coupon(request):
    if request.method == 'POST':
        code = request.POST.get("code")
        discount_price = request.POST.get("discount_price")
        min_cart_value = request.POST.get("min_cart_value")
        max_discount = request.POST.get("max_discount")
        max_usage_count = request.POST.get("max_usage_count")

        try:
            discount_price = float(discount_price)
            if discount_price < 1 or discount_price > 100:
                raise ValueError("Discount price must be between 1 and 100.")
        except (ValueError, TypeError):
            messages.error(request, 'Invalid discount price. Please enter a value between 1 and 100.')
            return render(request, 'adminpage/add_coupon.html')

        coupon = Coupon(
            code=code,
            discount_price=discount_price,
            minimum_cart_value = min_cart_value,
            maximum_discount = max_discount,
            maximum_usage_count = max_usage_count
        )
        coupon.save()
        return redirect('adminpage:coupon_management')

    return render(request, 'adminpage/add_coupon.html')

@user_passes_test(is_superuser, login_url='home:signin') 
def activate_deactivate_coupon(request,id):
    coupon=Coupon.objects.get(id=id)
    if coupon.is_active==True:
        coupon.is_active=False
    else:
        coupon.is_active=True
    coupon.save()
    return redirect('adminpage:coupon_management')

@user_passes_test(is_superuser, login_url='home:signin')
def coupon_update(request, id):
    coupon = Coupon.objects.get(id=id)

    if request.method == 'POST':
        code = request.POST.get("code")
        discount_price = request.POST.get("discount_price")
        min_cart_value = request.POST.get("min_cart_value")
        max_discount = request.POST.get("max_discount")
        max_usage_count = request.POST.get("max_usage_count")

        try:
            if discount_price:
                discount_price = float(discount_price)
                if discount_price < 1 or discount_price > 100:
                    raise ValueError("Discount price must be between 1 and 100.")
        except (ValueError, TypeError):
            messages.error(request, 'Invalid discount price. Please enter a value between 1 and 100.')
            return render(request, 'adminpage/coupon_update.html', {'cu': coupon})

        if code:
            coupon.code = code
        if discount_price:
            coupon.discount_price = discount_price
        if min_cart_value:
            coupon.minimum_cart_value = min_cart_value
        if max_discount:
            coupon.maximum_discount = max_discount
        if max_usage_count:
            coupon.maximum_usage_count = max_usage_count
        

        coupon.save()
        return redirect('adminpage:coupon_management')

    else:
        return render(request, 'adminpage/coupon_update.html', {'cu': coupon})

@user_passes_test(is_superuser, login_url='home:signin')
def edit_varient(request,id):
    variant = get_object_or_404(ProductVariant, id=id)
    print(variant)
    if request.method == 'POST':
        # Assuming your form fields are 'price' and 'quantity'
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')

        # Update the variant fields
        variant.price = price
        variant.quantity = quantity

        variant.save()
        return redirect('adminpage:product_management')
    return render(request,'adminpage/edit_varient.html',{'varient':variant})

@user_passes_test(is_superuser, login_url='home:signin')
def order_details(request,id):
    user_order = get_object_or_404(OrderDetails, id=id)
    return render(request,'adminpage/order_details.html',{'order':user_order})

@user_passes_test(is_superuser, login_url='home:signin')
def banner(request):
    banner = Banner.objects.all()
    return render(request,'adminpage/banner.html',{'banner':banner})

@user_passes_test(is_superuser, login_url='home:signin')
def add_banner(request):
    if request.method == 'POST':
        # Extract product details
        title = request.POST.get('title').strip()
        description = request.POST.get('description').strip()
        

        # Validations for title and description
        if not title or not description:
            messages.error(request, 'Title and description cannot be empty or spaces.')
            return redirect('adminpage:add_banner')  # Updated redirect URL

        # Handle main image upload
        main_image = request.FILES.get('image') if 'image' in request.FILES else None

        

        # Get an existing product or create a new one
        Banner.objects.get_or_create(
            title=title,
            defaults={
                'description': description,
                'image': main_image,
            }
        )

       

        return redirect('adminpage:banner')
    return render(request,'adminpage/add_banner.html')

@user_passes_test(is_superuser, login_url='home:signin')
def activate_deactivate_banner(request,id=id):
    if request.method=='POST':
        ban=Banner.objects.get(id=id)
        if ban.is_active==True:
            ban.is_active=False
        else:
            ban.is_active=True
        ban.save()
    return redirect('adminpage:banner')

@user_passes_test(is_superuser, login_url='home:signin')
def edit_banner(request,id):
    # Get the product instance using the product_id
    banner = get_object_or_404(Banner, id=id)
    if request.method == 'POST':
        # Process the form submission
        title = request.POST.get('title')
        description = request.POST.get('description')
        image = request.FILES.get('image') if 'image' in request.FILES else None
        # Update the product fields
        banner.title = title
        banner.description = description

        if image:
            banner.image = image

        banner.save()    

        return redirect('adminpage:banner')  # Redirect to the provider panel or any other desired page
    return render(request, 'adminpage/edit_banner.html',{'banner':banner})

@user_passes_test(is_superuser, login_url='home:signin')
def transactions(request):
    orders = OrderDetails.objects.all().order_by('-id')
    paginator = Paginator(orders, 5)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)
    return render(request, 'adminpage/transactions.html', {'orders': orders_page})

@user_passes_test(is_superuser, login_url='home:signin')
def export_order_details_to_excel(request):
    # Retrieve order details queryset
    orders = OrderDetails.objects.all()

    # Create a new Excel workbook and get the active sheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Order ID', 'User', 'Order Time', 'User Mobile', 'User Address',
               'Product', 'Product Variant', 'Product Quantity', 'Product Price',
               'Payment Type', 'Order Status', 'Date']

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Add data to the worksheet
    for row_num, order in enumerate(orders, 2):
        ws.cell(row=row_num, column=1, value=order.id)
        ws.cell(row=row_num, column=2, value=order.user.username)
        ws.cell(row=row_num, column=3, value=timezone.localtime(order.order_time).replace(tzinfo=None) if order.order_time else None)
        ws.cell(row=row_num, column=4, value=order.user_mobile.mobile_number if order.user_mobile else '')
        ws.cell(row=row_num, column=5, value=order.user_address)
        ws.cell(row=row_num, column=6, value=order.product.title)
        ws.cell(row=row_num, column=7, value=order.product_variant.size)
        ws.cell(row=row_num, column=8, value=order.product_quantity)
        ws.cell(row=row_num, column=9, value=float(order.product_price))  # Convert Decimal to float for Excel
        ws.cell(row=row_num, column=10, value=order.payment.payment_type if order.payment else '')
        ws.cell(row=row_num, column=11, value=order.order_status)
        ws.cell(row=row_num, column=12, value=order.date.strftime('%Y-%m-%d') if order.date else None)

    # Create a response object with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url='home:signin')
def offer_management(request):
    offer=ProductOffer.objects.all()
    return render(request,'adminpage/offer_management.html',{'offer':offer})

@user_passes_test(is_superuser, login_url='home:signin')
def add_offer(request):
    if request.method == 'POST':
        product_title = request.POST.get('product')
        discount_amount = request.POST.get('discount_amount')
        duration_minutes = request.POST.get('duration_minutes')  # Get the duration in minutes

        # Retrieve the Product instance based on the selected title
        product = Product.objects.get(title=product_title)

        # Create a ProductOffer instance and save it to the database
        offer = ProductOffer(product=product, discount_amount=discount_amount)
        offer.save()

        # Redirect to a success page or render a success template
        return redirect('adminpage:offer_management')  # Replace 'success-page-url' with the actual URL
    product = Product.objects.all()
    return render(request, 'adminpage/add_offer.html',{'product':product})

@user_passes_test(is_superuser, login_url='home:signin')
def delete_offer(request, id):
    # Get the offer object by its ID or return a 404 error if not found
    offer = get_object_or_404(ProductOffer, id=id)

    # Delete the offer
    offer.delete()

    # Redirect or render the response as needed
    # For example, rendering the offer management page again
    offers = ProductOffer.objects.all()
    return redirect('adminpage:offer_management')

@user_passes_test(is_superuser, login_url='home:signin')
def search_users(request):
    query = request.GET.get('q')
    users = User.objects.exclude(is_superuser=True).order_by('id')
    
    if query:
        users = users.filter(
            Q(username__icontains=query) | 
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number)
    
    return render(request, 'adminpage/user_management.html', {'users': users_page})
